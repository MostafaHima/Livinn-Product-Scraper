import openpyxl


class ExcelExporter:
    def __init__(self, file_name, data):
        self.file_name = file_name  # Excel file name
        self.data = data  # Data to be exported
        self.workbook = openpyxl.Workbook()  # Create a new Excel workbook

    def create_sheets(self):
        """Create sheets and fill them with data."""
        print("📄 Starting sheet creation process...")
        # Remove the default sheet if it exists
        if "Sheet" in self.workbook.sheetnames:
            default_sheet = self.workbook["Sheet"]
            self.workbook.remove(default_sheet)
            print("🗑️ Default sheet removed.")

        # Create new sheets
        for index, sheet_data in enumerate(self.data):
            sheet_name = f"Sheet {index + 1}"  # Default sheet name
            # Assign specific names based on categories
            sheet_name = {
                0: "Food",
                1: "Health",
                2: "Fashion",
                3: "Home",
                4: "Kids",
            }.get(index, sheet_name)

            sheet = self.workbook.create_sheet(title=sheet_name)
            print(f"📄 Sheet '{sheet_name}' created.")
            self.fill_sheet(sheet, sheet_data)

    def fill_sheet(self, sheet, sheet_data):
        """Fill a specific sheet with data."""
        print(f"✍️ Filling data into sheet '{sheet.title}'...")
        # Check if data exists
        if not sheet_data:
            print(f"⚠️ No data found for sheet '{sheet.title}'. Skipping...")
            return

        # Add column headers
        headers = list(sheet_data[0].keys())
        sheet.append(headers)
        print(f"🗂️ Headers added: {headers}")

        # Add data rows
        for row_data in sheet_data:
            row = [row_data.get(header, "") for header in headers]
            sheet.append(row)
        print(f"✅ Data filled into sheet '{sheet.title}' successfully!")

    def save_file(self):
        """Save the file."""
        try:
            self.workbook.save(self.file_name)
            print(f"💾 Excel file '{self.file_name}' has been created successfully!")
        except Exception as e:
            print(f"❌ Error saving file: {e}")
